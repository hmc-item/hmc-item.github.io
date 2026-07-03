#!/usr/bin/env python3
# 조·역량 세팅 스크립트 — '조_역량_세팅_템플릿.xlsx'를 읽어 Supabase teams/competencies에 등록.
# 사용:
#   python seed_teams_comps.py <템플릿.xlsx>            # DRY-RUN(미등록, 계획만 출력)
#   python seed_teams_comps.py <템플릿.xlsx> --commit   # 실제 등록
#   옵션 --wipe : 등록 전 기존 teams/competencies 전부 삭제(테스트 데이터 교체용)
import sys, json, time, urllib.request, urllib.parse
import openpyxl

URL = "https://ozykvjuktmfyxdmudooh.supabase.co"
KEY = "sb_publishable_Ht9q_kAOXQHPDCjTu_7oqQ_64Jnr2C3"
H = {"apikey": KEY, "Authorization": "Bearer " + KEY, "Content-Type": "application/json"}


def req(method, path, body=None):
    url = URL + "/rest/v1/" + path
    data = json.dumps(body).encode("utf-8") if body is not None else None
    r = urllib.request.Request(url, data=data, headers=H, method=method)
    with urllib.request.urlopen(r) as resp:
        raw = resp.read().decode("utf-8")
        return resp.status, (json.loads(raw) if raw.strip() else None)


def read_template(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    teams = []
    ws = wb["조"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = (str(row[0]).strip() if row and row[0] else "")
        if not name:
            continue
        cls = None
        if len(row) > 1 and row[1] not in (None, ""):
            try:
                cls = int(str(row[1]).strip())
            except ValueError:
                cls = None
        teams.append({"team_name": name, "class_no": cls})
    comps = []
    ws = wb["역량"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = (str(row[0]).strip() if row and row[0] else "")
        if not name:
            continue
        team = (str(row[1]).strip() if len(row) > 1 and row[1] else "")
        target = 50
        if len(row) > 2 and row[2] not in (None, ""):
            try:
                target = int(str(row[2]).strip())
            except ValueError:
                target = 50
        cat = (str(row[3]).strip() if len(row) > 3 and row[3] else "")
        comps.append({"comp_name": name, "team": team, "target_count": target, "category": cat})
    return teams, comps


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    flags = [a for a in sys.argv[1:] if a.startswith("--")]
    if not args:
        print("사용: python seed_teams_comps.py <템플릿.xlsx> [--commit] [--wipe]")
        return
    path = args[0]
    commit = "--commit" in flags
    wipe = "--wipe" in flags

    teams, comps = read_template(path)
    print(f"[읽음] 조 {len(teams)}개 / 역량 {len(comps)}개")

    # 담당조 이름 → team 검증
    team_names = {t["team_name"] for t in teams}
    bad = [c for c in comps if c["team"] and c["team"] not in team_names]
    if bad:
        print("⚠️ [경고] 담당조가 [조] 시트에 없는 역량:")
        for c in bad:
            print(f"    - {c['comp_name']} → 담당조 '{c['team']}' 없음")

    if not commit:
        print("\n=== DRY-RUN (미등록) ===")
        for t in teams:
            print(f"  조: {t['team_name']} (분반 {t['class_no']})")
        for c in comps:
            print(f"  역량: {c['comp_name']} → {c['team']} / 목표 {c['target_count']} / 분류 {c['category'] or '-'}")
        print("\n실제 등록하려면 --commit 를 붙이세요.")
        return

    ts = int(time.time())
    if wipe:
        print("[wipe] 기존 competencies/teams 삭제 중...")
        req("DELETE", "competencies?comp_id=neq.__none__")
        req("DELETE", "teams?team_id=neq.__none__")

    name_to_id = {}
    for i, t in enumerate(teams):
        tid = f"t_{ts}_{i}"
        body = {"id": tid, "team_id": tid, "team_name": t["team_name"], "class_no": t["class_no"]}
        st, _ = req("POST", "teams", body)
        name_to_id[t["team_name"]] = tid
        print(f"  [team {st}] {t['team_name']} → {tid}")

    for i, c in enumerate(comps):
        cid = f"c_{ts}_{i}"
        body = {"id": cid, "comp_id": cid, "comp_name": c["comp_name"],
                "category": c["category"] or None, "description": None,
                "team_id": name_to_id.get(c["team"]), "target_count": c["target_count"],
                "order_index": i}
        st, _ = req("POST", "competencies", body)
        print(f"  [comp {st}] {c['comp_name']} → {c['team']} ({name_to_id.get(c['team'])})")

    print("완료.")


if __name__ == "__main__":
    main()
